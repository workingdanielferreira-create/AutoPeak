"""
bridge.py
=========
Local-file I/O ONLY. No network code lives here — that's the entire point of
the zero-IT design. Python writes/reads plain files in a folder that happens
to be a OneDrive-synced SharePoint document library ("the Drop folder").
Power Automate flows, running on the manager's own already-permitted Office
365 connections, watch that same folder and do the actual cloud calendar work.

Supports BOTH outcomes of the §7 validation test via cfg["write_path"]:

  "primary"  - one shared busy_<month>.json / results_<month>.json, produced
               by Flow A / Flow B (manager's connection writes to coach
               calendars that were shared with them as "Can edit").
  "fallback" - one busy_<month>_<alias>.json / results_<month>_<alias>.json
               PER COACH, each produced by that coach's own imported flow
               (each coach owns the write to their own calendar).

`proposals_<month>.json` is written ONCE regardless of path — in fallback,
every coach's flow watches the same file and each just filters to its own
`owner` before acting, so the contract doesn't fork there.

Nothing here ever calls Outlook directly — that stays in outlook_client.py
(manager's own calendar, via COM) and inside the Power Automate flows
(coach calendars, via their Office 365 Outlook connector).
"""

from __future__ import annotations
import json
import os
import re
import time
from datetime import datetime

try:
    import openpyxl  # type: ignore
    _HAVE_OPENPYXL = True
except Exception:
    _HAVE_OPENPYXL = False


class BridgeError(RuntimeError):
    pass


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _alias(email: str) -> str:
    """Turn an email into a filesystem-safe alias for per-coach filenames."""
    local = (email or "").split("@")[0].strip().lower()
    return re.sub(r"[^a-z0-9_-]+", "_", local) or "unknown"


def _path(drop_folder: str, filename: str) -> str:
    if not drop_folder:
        raise BridgeError(
            "config.json has no 'drop_folder' set — point it at your "
            "OneDrive-synced AutoPeak Drop folder first."
        )
    return os.path.join(drop_folder, filename)


def _write_json(path: str, data) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    os.replace(tmp, path)  # atomic-ish on the local filesystem


def _read_json(path: str):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _iso(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%dT%H:%M:%S")


def _parse_iso(s: str) -> datetime:
    return datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S")


# --------------------------------------------------------------------------- #
# 1. Coach/agent mapping  (CoachAgentMapping.xlsx -> agents_by_coach)
# --------------------------------------------------------------------------- #
def read_mapping(xlsx_path: str) -> dict[str, list[str]]:
    """Read the shared CoachAgentMapping.xlsx (columns: CoachEmail, AgentEmail)
    and return {coach_email: [agent_email, ...]}. Coaches sharing the sheet
    just add/remove rows; nothing else needs to change for them.
    """
    if not _HAVE_OPENPYXL:
        raise BridgeError(
            "openpyxl isn't installed. Run: python -m pip install openpyxl --break-system-packages"
        )
    if not os.path.exists(xlsx_path):
        raise BridgeError(f"Mapping file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        coach_i = header.index("coachemail")
        agent_i = header.index("agentemail")
    except ValueError:
        raise BridgeError(
            "CoachAgentMapping.xlsx must have header row 'CoachEmail', 'AgentEmail'."
        )

    out: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2):
        coach = str(row[coach_i].value or "").strip()
        agent = str(row[agent_i].value or "").strip()
        if not coach or not agent:
            continue
        out.setdefault(coach, [])
        if agent not in out[coach]:
            out[coach].append(agent)
    return out


def merge_mapping_into_cfg(cfg: dict) -> dict:
    """Load the mapping sheet (if configured) and fold it into cfg in place:
    fills agents_by_coach, and adds any coach found in the sheet that wasn't
    already listed in cfg['coaches']. Safe to call every run."""
    xlsx_path = cfg.get("agent_mapping_xlsx")
    if not xlsx_path:
        return cfg
    full_path = xlsx_path
    if not os.path.isabs(full_path) and cfg.get("drop_folder"):
        full_path = os.path.join(cfg["drop_folder"], os.path.basename(xlsx_path))
    mapping = read_mapping(full_path)
    cfg["agents_by_coach"] = mapping
    existing = list(cfg.get("coaches", []))
    for coach in mapping:
        if coach not in existing:
            existing.append(coach)
    cfg["coaches"] = existing
    return cfg


# --------------------------------------------------------------------------- #
# 2. request_<month>.json   (tool -> Flow A)
# --------------------------------------------------------------------------- #
def write_request(cfg: dict, owners: list[str]) -> str:
    path = _path(cfg["drop_folder"], f"request_{cfg['month']}.json")
    _write_json(path, {"month": cfg["month"], "owners": owners})
    return path


# --------------------------------------------------------------------------- #
# 3. busy_<month>.json (primary) / busy_<month>_<alias>.json (fallback)
#    Flow A / per-coach flow -> tool
# --------------------------------------------------------------------------- #
def busy_ready(cfg: dict, coaches: list[str]) -> bool:
    """True once every expected busy file has shown up."""
    try:
        read_busy(cfg, coaches)
        return True
    except BridgeError:
        return False


def read_busy(cfg: dict, coaches: list[str]):
    """Returns {owner_email: [(start_dt, end_dt), ...]} for every coach.
    Raises BridgeError if a required file hasn't appeared yet — callers
    should treat that as 'still waiting on the flow', not a hard failure.
    Does NOT include the manager — manager busy comes from outlook_client
    (COM) directly, since that's local and instant."""
    out: dict[str, list[tuple[datetime, datetime]]] = {}
    drop = cfg["drop_folder"]
    write_path = cfg.get("write_path", "primary")

    if write_path == "fallback":
        for coach in coaches:
            p = _path(drop, f"busy_{cfg['month']}_{_alias(coach)}.json")
            if not os.path.exists(p):
                raise BridgeError(f"Still waiting on {os.path.basename(p)}")
            data = _read_json(p)
            out[coach] = [(_parse_iso(s), _parse_iso(e)) for s, e in data]
    else:  # primary
        p = _path(drop, f"busy_{cfg['month']}.json")
        if not os.path.exists(p):
            raise BridgeError(f"Still waiting on {os.path.basename(p)}")
        data = _read_json(p)
        for coach in coaches:
            pairs = data.get(coach, [])
            out[coach] = [(_parse_iso(s), _parse_iso(e)) for s, e in pairs]
    return out


def wait_for_busy(cfg: dict, coaches: list[str], timeout_seconds: int = 60,
                  poll_seconds: int = 3):
    """Poll for the busy file(s) up to timeout_seconds. Returns the busy dict,
    or raises BridgeError if it never showed up — that almost always means
    the flow hasn't run yet (check its run history in Power Automate)."""
    deadline = time.time() + timeout_seconds
    last_err = None
    while time.time() < deadline:
        try:
            return read_busy(cfg, coaches)
        except BridgeError as e:
            last_err = e
            time.sleep(poll_seconds)
    raise BridgeError(
        f"Timed out waiting for coach availability ({last_err}). "
        "Check the flow's run history in Power Automate, then try again."
    )


# --------------------------------------------------------------------------- #
# 4. proposals_<month>.json   (tool -> Flow B, same file in both paths)
# --------------------------------------------------------------------------- #
def write_proposals(cfg: dict, proposals: list) -> tuple[str, list[str]]:
    """proposals: list of scheduler.Proposal objects, ALREADY filtered to only
    the coach-owned ones (manager-owned go through outlook_client.commit
    instead — they never touch the bridge). Returns (path, ids_written)."""
    rows = []
    ids = []
    for i, p in enumerate(proposals, start=1):
        pid = f"p{i:03d}"
        ids.append(pid)
        rows.append({
            "id": pid,
            "owner": p.owner,
            "name": p.name,
            "category": p.category,
            "schedule_type": p.schedule_type,
            "start": _iso(p.start),
            "end": _iso(p.end),
            "attendee": p.attendees[0] if p.attendees else "",
            "status": "Pending",
        })
    path = _path(cfg["drop_folder"], f"proposals_{cfg['month']}.json")
    _write_json(path, rows)
    return path, ids


# --------------------------------------------------------------------------- #
# 5. results_<month>.json (primary) / results_<month>_<alias>.json (fallback)
#    Flow B / per-coach flow -> tool
# --------------------------------------------------------------------------- #
def read_results(cfg: dict, coaches: list[str]) -> list[dict]:
    """Returns a flat list of {id, status, error?} across however many result
    files exist right now (partial results are fine — callers can re-poll)."""
    drop = cfg["drop_folder"]
    write_path = cfg.get("write_path", "primary")
    out: list[dict] = []

    if write_path == "fallback":
        for coach in coaches:
            p = _path(drop, f"results_{cfg['month']}_{_alias(coach)}.json")
            if os.path.exists(p):
                out.extend(_read_json(p))
    else:
        p = _path(drop, f"results_{cfg['month']}.json")
        if os.path.exists(p):
            out.extend(_read_json(p))
    return out


def wait_for_results(cfg: dict, coaches: list[str], expected_ids: list[str],
                     timeout_seconds: int = 60, poll_seconds: int = 3) -> list[dict]:
    """Poll until every expected id has a result, or timeout (returns whatever
    arrived so far — the caller should report partial completion clearly)."""
    deadline = time.time() + timeout_seconds
    seen: dict[str, dict] = {}
    while time.time() < deadline:
        for r in read_results(cfg, coaches):
            seen[r["id"]] = r
        if all(i in seen for i in expected_ids):
            break
        time.sleep(poll_seconds)
    return list(seen.values())
